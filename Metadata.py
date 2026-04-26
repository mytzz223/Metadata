import discord
from discord.ext import commands
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

# ====== CONFIG ======
TOKEN = "MTQ5NTA4MDUxODM1MjYzODExNA.GXTEjh.5CrjgFCcimzR97yhc-fCRjFTWHjtBPmcf8pAK4"
LOG_CHANNEL_ID = 1497938392212701245  # 🔁 PUT YOUR CHANNEL ID

FILE_NAME = "users.xlsx"
LOG_TXT = "logs.txt"

# ====== INTENTS ======
intents = discord.Intents.default()
intents.members = True

bot = commands.Bot(command_prefix="!", intents=intents)

# ====== EXCEL SETUP ======
def setup_excel():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append([
            "User ID", "Username", "Account Created",
            "Joined At", "Left At", "Avatar URL"
        ])
        wb.save(FILE_NAME)

# ====== ADD USER TO EXCEL ======
def add_user(member):
    wb = load_workbook(FILE_NAME)
    ws = wb.active

    ws.append([
        str(member.id),  # ✅ FIXED (no scientific notation)
        str(member),
        member.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        member.joined_at.strftime("%Y-%m-%d %H:%M:%S") if member.joined_at else "N/A",
        "Still in server",
        member.display_avatar.url
    ])

    wb.save(FILE_NAME)

# ====== MARK USER LEFT ======
def mark_left(member):
    wb = load_workbook(FILE_NAME)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(member.id) and row[4].value == "Still in server":
            row[4].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            break

    wb.save(FILE_NAME)

# ====== TXT LOGGING ======
def write_log(text):
    with open(LOG_TXT, "a", encoding="utf-8") as f:
        f.write(text + "\n")

# ====== READY ======
@bot.event
async def on_ready():
    setup_excel()
    print(f"✅ Logged in as {bot.user}")

# ====== MEMBER JOIN ======
@bot.event
async def on_member_join(member):
    add_user(member)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # TXT LOG
    write_log(f"[JOIN] Username: {member} | ID: {member.id} | Time: {now}")

    # DISCORD LOG
    channel = bot.get_channel(LOG_CHANNEL_ID)
    if channel:
        embed = discord.Embed(
            title="🟢 Member Joined",
            color=discord.Color.green(),
            timestamp=datetime.utcnow()
        )

        embed.add_field(name="Username", value=str(member), inline=False)
        embed.add_field(name="User ID", value=member.id, inline=False)
        embed.add_field(
            name="Account Created",
            value=member.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            inline=False
        )

        # PFP
        embed.set_thumbnail(url=member.display_avatar.url)
        embed.set_image(url=member.display_avatar.url)

        await channel.send(embed=embed)

# ====== MEMBER LEAVE ======
@bot.event
async def on_member_remove(member):
    mark_left(member)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # TXT LOG
    write_log(f"[LEAVE] Username: {member} | ID: {member.id} | Time: {now}")

    # DISCORD LOG
    channel = bot.get_channel(LOG_CHANNEL_ID)
    if channel:
        embed = discord.Embed(
            title="🔴 Member Left",
            color=discord.Color.red(),
            timestamp=datetime.utcnow()
        )

        embed.add_field(name="Username", value=str(member), inline=False)
        embed.add_field(name="User ID", value=member.id, inline=False)

        # PFP
        embed.set_thumbnail(url=member.display_avatar.url)
        embed.set_image(url=member.display_avatar.url)

        await channel.send(embed=embed)

# ====== RUN ======
bot.run(TOKEN)
